
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
import pandas as pd

sample_urls = ["https://vantage-mortgages.co.uk/","https://wy-money.co.uk/","https://mortgagefactoryltd.com/","https://www.domusfinancial.co.uk/"]
chrome_options = Options()
chrome_options.add_argument("--headless")
driver = webdriver.Chrome(options=chrome_options)


from data_clean import URLStatusChecker

checker = URLStatusChecker('test_accounts.csv', n_rows=575)
checker.test_urls('url')
cleaned_df = checker.get_dataframe()

def get_soup(url):
    driver.get(url)
    driver.implicitly_wait(4)  # wait for 4 seconds for the page to load
    web_page = driver.page_source
    soup = BeautifulSoup(web_page, 'html.parser')
    return soup


def social_media_links(page_soup):

    social_media_sites = ["facebook","instagram","twitter"]
    social_dict = {}

    for item in social_media_sites:
        try:
            web_element = page_soup.find(name="a", href=lambda href: href and f"{item}.com" in href)
            web_element_link = web_element['href']
            social_dict[item] = web_element_link
        except:
            social_dict[item] = "no link found"

    return social_dict


def about_us_page(page_soup, base_url):
    # Separate the keywords related to "team" and others for prioritization
    team_keywords = ["meet-the-team", "meet-team", "our-people", "team"]
    other_keywords = ["about", "our-story", "our-process"]

    # Combine the keywords for a single iteration
    combined_keywords = team_keywords + other_keywords

    for keyword in combined_keywords:
        try:
            link = page_soup.find('a', href=lambda href: href and keyword.lower() in href.lower())
            if link and link['href']:
                about_link = link['href']
                # Check if the link is relative
                if not about_link.startswith('http'):
                    # Construct the full URL
                    about_link = base_url + about_link
                return about_link
        except:
            continue

    return "no link found"


copy_table = cleaned_df.copy()
good_links = copy_table[copy_table["status"] != "broken website"]
urls = good_links["url_clean"]

def export_dict(url_list):
    entries = []
    for url in url_list:
        print(f"trying URL: {url}")
        page_soup = get_soup(url=url)
        socials = social_media_links(page_soup=page_soup)
        abouts = about_us_page(page_soup=page_soup, base_url=url)
        socials["about"] = abouts
        socials["url_clean"] = url
        entries.append(socials)
    return entries

new_df = pd.DataFrame(export_dict(urls))
another_new_df = pd.merge(cleaned_df, new_df, on="url_clean", how="left")

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

# Assuming another_new_df is your final DataFrame

# Define your output Excel file
output_file = 'test_output.xlsx'

# Create an Excel writer using openpyxl
writer = pd.ExcelWriter(output_file, engine='openpyxl')

# Write your DataFrame to an Excel file
another_new_df.to_excel(writer, index=False, sheet_name='Sheet1')

# Load the workbook and select the first sheet
workbook = writer.book
worksheet = writer.sheets['Sheet1']

# Iterate over the rows and format the last 5 columns as hyperlinks
for row in worksheet.iter_rows(min_row=2, max_col=worksheet.max_column, max_row=worksheet.max_row):
    selected_cells = [row[4]] + list(row[-5:])
    for cell in selected_cells:
        if cell.value != 'no link found':
            cell.value = f'=HYPERLINK("{cell.value}", "{cell.value}")'
            cell.font = Font(color='0000FF', underline='single')

# Save the workbook
writer.close()




