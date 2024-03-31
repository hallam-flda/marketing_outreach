from data_clean import URLStatusChecker
from web_elements import WebScraper
from gpt_commenter import GPTCommenter
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

checker = URLStatusChecker('test_accounts.csv',n_rows=574)
checker.test_urls('url')
full_df = checker.get_dataframe()

full_df_2 = full_df.copy()
okay_link_df = full_df_2[full_df_2["status"] != 'broken website']

web_scraper = WebScraper(okay_link_df)
scraped_data = web_scraper.scrape_data()

scraped_data_2 = scraped_data.copy()

commenter = GPTCommenter(input_df=scraped_data_2)
commenter.process_urls()

abc = commenter.result_df

print(abc.head())
print(abc.info())
print(full_df.info())

required_col_df = full_df_2[["first_name","company_name","email","status","url_clean"]]

merged_df = pd.merge(required_col_df,abc, on="url_clean", how="left")
merged_df.info()


# Define your output Excel file
output_file = 'full_output_575.xlsx'

# Create an Excel writer using openpyxl
writer = pd.ExcelWriter(output_file, engine='openpyxl')

# Write your DataFrame to an Excel file
merged_df.to_excel(writer, index=False, sheet_name='Sheet1')

# Load the workbook and select the first sheet
workbook = writer.book
worksheet = writer.sheets['Sheet1']

# Iterate over the rows and format the last 5 columns as hyperlinks
for row in worksheet.iter_rows(min_row=2, max_col=worksheet.max_column, max_row=worksheet.max_row):
    selected_cells = row[4:11]
    for cell in selected_cells:
        if cell.value != 'no link found':
            cell.value = f'=HYPERLINK("{cell.value}", "{cell.value}")'
            cell.font = Font(color='0000FF', underline='single')

# Save the workbook
writer.close()
